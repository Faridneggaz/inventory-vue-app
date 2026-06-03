"""
Inventory API Module
Provides server-side logic for fetching item details with stock and pricing information.
"""
import frappe
from frappe.model.document import Document
from frappe import _
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

@frappe.whitelist()
def get_items_with_details(item_group=None, warehouse=None, buying_price_list=None, selling_price_list=None):
    """
    Retrieves a comprehensive list of items including current stock (Bin),
    barcodes, and pricing from specified price lists.

    Args:
        item_group (str, optional): Filter by Item Group.
        warehouse (str, optional): Filter stock levels for this warehouse.
        buying_price_list (str, optional): Price list for purchase rates.
        selling_price_list (str, optional): Price list for sales rates.

    Returns:
        list[dict]: List of item dictionaries with code, name, barcode, qty, and rates.
    """
    query = """
        SELECT
            item.name as item_code,
            item.item_name as item_name,
            (SELECT barcode FROM `tabItem Barcode` WHERE parent = item.name LIMIT 1) as barcode,
            IFNULL(bin.actual_qty, 0) as qty,
            IFNULL(bp.price_list_rate, 0) as buying_rate,
            IFNULL(sp.price_list_rate, 0) as selling_rate
        FROM
            `tabItem` item
        LEFT JOIN
            `tabBin` bin ON bin.item_code = item.name AND bin.warehouse = %(warehouse)s
        LEFT JOIN
            `tabItem Price` bp ON bp.item_code = item.name AND bp.price_list = %(buying_list)s
        LEFT JOIN
            `tabItem Price` sp ON sp.item_code = item.name AND sp.price_list = %(selling_list)s
        WHERE
            item.disabled = 0
    """
    
    values = {
        "warehouse": warehouse,
        "buying_list": buying_price_list,
        "selling_list": selling_price_list
    }

    if item_group:
        query += " AND item.item_group = %(item_group)s"
        values["item_group"] = item_group

    return frappe.db.sql(query, values, as_dict=True)


@frappe.whitelist()
def export_inventory_to_excel(inventory_name):
    """
    Exports inventory items to Excel format with styling.

    Args:
        inventory_name (str): The name of the FSM Inventory document.

    Returns:
        dict: Contains the file URL and download information.
    """
    # Fetch the inventory document
    inventory = frappe.get_doc('FSM Inventory', inventory_name)

    if not inventory or not inventory.fsm_inventory_item:
        frappe.throw(_('No items to export'))

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Inventory Information Sheet
    inv_ws = wb.active
    inv_ws.title = 'Inventory Info'

    # Add inventory information data
    inv_data = [
        ['INVENTORY INFORMATION'],
        [''],
        ['Field', 'Value'],
        ['Name', inventory.name or ''],
        ['Reference', inventory.inventory_reference or ''],
        ['Warehouse', inventory.warehouse or ''],
        ['Family', inventory.group or ''],
        ['Starting Date', inventory.starting_date or ''],
        ['End Date', inventory.end_date or ''],
        ['Buying Price List', inventory.buying_price_list or ''],
        ['Selling Price List', inventory.selling_price_list or ''],
        ['Total Expected Qty', inventory.total_expected_qty or 0],
        ['Total Counted Qty', inventory.total_counted_qty or 0],
        ['Total Qty Offset', inventory.total_qty_offset or 0],
        ['Status', 'Submitted' if inventory.docstatus == 1 else 'Draft']
    ]

    for row in inv_data:
        inv_ws.append(row)

    # Style the inventory sheet
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for row in inv_ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = header_alignment

    for row in inv_ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font = header_font
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = header_alignment

    # Set column widths for inventory sheet
    inv_ws.column_dimensions['A'].width = 30
    inv_ws.column_dimensions['B'].width = 40

    # Items Sheet
    items_ws = wb.create_sheet('Items')

    # Add items data
    items_data = [
        ['INVENTORY ITEMS'],
        [''],
        ['Item Code', 'Item Name', 'Barcode', 'Expected Qty', 'Counted Qty', 'Offset', 'Buying Price', 'Selling Price', 'Scanned', 'Scanned At']
    ]

    for item in inventory.fsm_inventory_item:
        items_data.append([
            item.item_code or '',
            item.item_name or '',
            item.barcode or '',
            item.expected_qty or 0,
            item.counted_qty or 0,
            item.qty_offset or 0,
            item.buying_price or 0,
            item.selling_price or 0,
            'Yes' if item.is_scanned else 'No',
            item.scanned_at or ''
        ])

    for row in items_data:
        items_ws.append(row)

    # Style the items sheet
    for row in items_ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True, size=14, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = header_alignment

    for row in items_ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.font = header_font
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = header_alignment

    # Set column widths for items sheet
    items_ws.column_dimensions['A'].width = 20
    items_ws.column_dimensions['B'].width = 30
    items_ws.column_dimensions['C'].width = 18
    items_ws.column_dimensions['D'].width = 14
    items_ws.column_dimensions['E'].width = 14
    items_ws.column_dimensions['F'].width = 12
    items_ws.column_dimensions['G'].width = 14
    items_ws.column_dimensions['H'].width = 14
    items_ws.column_dimensions['I'].width = 10
    items_ws.column_dimensions['J'].width = 22

    # Save to a BytesIO buffer
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    # Generate file name
    file_name = f"{inventory.name or 'inventory'}_export_{frappe.utils.nowdate().replace('-', '')}.xlsx"

    # Save the file to Frappe's file system
    file_doc = frappe.get_doc({
        'doctype': 'File',
        'file_name': file_name,
        'content': file_buffer.getvalue(),
        'is_private': 0
    })
    file_doc.save()

    return {
        'file_url': file_doc.file_url,
        'file_name': file_name
    }












